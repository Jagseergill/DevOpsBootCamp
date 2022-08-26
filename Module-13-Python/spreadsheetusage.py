import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
product_under_10 = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculation number of products per supplier

    if supplier_name in product_per_supplier:
        current_num_product = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_product + 1
    else:
        print("Adding a new supplier")
        product_per_supplier[supplier_name] = 1

    # Calculate total value per supplier

    if supplier_name in total_value_per_supplier:
        current_total_Value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_Value + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Product Under 10

    if inventory < 10:
        product_under_10[product_num] = inventory
    # total value save in new column
    inventory_price.value = inventory * price
inv_file.save("Inventory-with-total-value.xlsx")
print(product_per_supplier)
print(total_value_per_supplier)
print(product_under_10)
