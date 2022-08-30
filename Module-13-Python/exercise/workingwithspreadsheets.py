from operator import itemgetter

import openpyxl

emp_file = openpyxl.load_workbook("employees.xlsx")
emp_list = emp_file["Sheet1"]

emp_list.delete_cols(3, 4)

employee_by_experience = []

for empl_row in range(2, emp_list.max_row + 1):
    empl_name = emp_list.cell(empl_row, 1).value
    empl_experience = emp_list.cell(empl_row, 2).value

    employee_by_experience.append({
        "name": empl_name,
        "experience": int(empl_experience)
    })


# sort the list of dictionaries by experience
new_list = sorted(employee_by_experience, key=itemgetter("experience"), reverse=True)


for empl_row in range(2, emp_list.max_row + 1):
    empl_name = emp_list.cell(empl_row, 1)
    empl_experience = emp_list.cell(empl_row, 2)

    # because the rows start from row 2, but index for our list starts at 0
    index = empl_row - 2
    employee = new_list[index]

    empl_name.value = employee["name"]
    empl_experience.value = employee["experience"]

emp_file.save("employees_sorted_by_experience.xlsx")
